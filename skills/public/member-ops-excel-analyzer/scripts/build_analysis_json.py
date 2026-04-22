#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


LEVEL_MAP = {1: "大众", 2: "青铜", 3: "白银", 4: "黄金", 5: "铂金", 6: "SVIP"}
LEVEL_ORDER = ["SVIP", "铂金", "黄金", "白银", "青铜", "大众"]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build member-ops analysis JSON from fixed template workbook.")
    parser.add_argument("--input", required=True, help="Path to the Excel workbook.")
    parser.add_argument("--output", required=True, help="Path to save analysis JSON.")
    return parser.parse_args()


def to_numeric(df: pd.DataFrame, cols: list[str]) -> None:
    for col in cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")


def simplify_store_type(raw: str) -> str:
    if raw == "专卖店":
        return "专卖店"
    if raw == "涂装服务中心":
        return "涂装服务中心"
    return "其他店型"


def control_store_type(raw: str) -> str:
    if raw == "专卖店":
        return "专卖店"
    if raw == "涂装服务中心":
        return "涂装服务中心"
    return "非可控店"


def read_fixed_area_tables(path: Path) -> tuple[list[list], list[list]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["扫码>=50游离会员"]

    jobs = []
    for row in range(3, 17):
        job = ws.cell(row, 8).value
        total = ws.cell(row, 9).value
        if job is None or total in (None, 0):
            continue
        jobs.append(
            [
                job,
                int(total),
                int(ws.cell(row, 10).value or 0),
                int(ws.cell(row, 11).value or 0),
                int(ws.cell(row, 12).value or 0),
                int(ws.cell(row, 13).value or 0),
            ]
        )

    total_row = ws.cell(35, 10).value
    if total_row is not None:
        jobs.append(
            [
                "合计",
                int(ws.cell(35, 10).value or 0),
                0,
                0,
                0,
                0,
            ]
        )

    areas = []
    for row in range(23, 36):
        city = ws.cell(row, 8).value
        area = ws.cell(row, 9).value
        if city is None and area is None:
            continue
        areas.append(
            [
                city or "",
                area or "",
                int(ws.cell(row, 10).value or 0),
                int(ws.cell(row, 11).value or 0),
                int(ws.cell(row, 12).value or 0),
                int(ws.cell(row, 13).value or 0),
                int(ws.cell(row, 14).value or 0),
                int(ws.cell(row, 15).value or 0),
                int(ws.cell(row, 16).value or 0),
                int(ws.cell(row, 17).value or 0),
                int(ws.cell(row, 18).value or 0),
                int(ws.cell(row, 19).value or 0),
                int(ws.cell(row, 20).value or 0),
                int(ws.cell(row, 21).value or 0),
                int(ws.cell(row, 22).value or 0),
                int(ws.cell(row, 23).value or 0),
                int(ws.cell(row, 24).value or 0),
            ]
        )
    return jobs, areas


def build_analysis(path: Path) -> dict:
    excel = str(path)
    member = pd.read_excel(excel, sheet_name="全部会员", header=1)
    store = pd.read_excel(excel, sheet_name="门店别", header=0, skiprows=[1])
    attached = pd.read_excel(excel, sheet_name="一物一码来源已挂靠会员分析", header=1, usecols="A:N")
    member_day_detail = pd.read_excel(excel, sheet_name="每月会员日参与名单", header=0)

    member_numeric_cols = [
        "等级",
        "性别",
        "年龄",
        "2025年门店消费金额（千元）",
        "2025年门店消费天数",
        "2025最近一次消费月",
        "2025年一物一码扫码包数",
        "2025年上门基检次数",
        "2025年参加会员日次数",
        "2025年会员商城购买金额（千元）",
        "2025年商城消费天数",
        "2025最近一次商城消费月",
        "SVIP拜访情况（2为已拜访）",
        "SVIP券获得数量",
        "SVIP券核销数量",
        "SVIP券连带消费金额（元）",
        "基检券获得数量",
        "基检券核销数量",
        "基检券连带消费金额（元）",
        "大奖中奖数量\n（汽车、电动车）",
        "普通实物中奖数量\n（不含车&色卡&辅材）",
        "色卡中奖数量",
        "辅材中奖数量",
        "券中奖数量",
        "辅材核销数量",
        "券核销数量",
        "券连带消费金额（元）",
    ]
    store_numeric_cols = [
        "累计注册人数（含挂靠）",
        "2025年门店消费人数",
        "2025年消费率",
        "2025年门店消费A价(未税 千元)",
        "一物一码来源挂靠人数",
        "其中：一物一码≥50包挂靠人数",
        "上门基检25年招募会员人数",
        "上门基检25年维护会员人数",
        "2025年联谊会签到人数",
        "SVIP人数",
        "2025年参加会员日抽奖人数",
        "2025年会员日有购买的人数",
        "2025年会员商城有购买的人数",
    ]
    attached_numeric_cols = list(attached.columns[4:])

    to_numeric(member, member_numeric_cols)
    to_numeric(store, store_numeric_cols)
    to_numeric(attached, attached_numeric_cols)

    member["等级名称"] = member["等级"].map(LEVEL_MAP)
    member["门店类型简"] = member["挂靠门店类型"].map(simplify_store_type)
    member["门店类型控"] = member["挂靠门店类型"].map(control_store_type)
    member["has_store_buy"] = member["2025年门店消费金额（千元）"].fillna(0) > 0
    member["has_scan"] = member["2025年一物一码扫码包数"].fillna(0) > 0
    member["has_member_day"] = member["2025年参加会员日次数"].fillna(0) > 0
    member["has_mall_buy"] = member["2025年会员商城购买金额（千元）"].fillna(0) > 0
    member["has_visit"] = member["2025年上门基检次数"].fillna(0) > 0
    member["has_union"] = member["has_store_buy"] | member["has_scan"]
    member["active_count"] = member[["has_store_buy", "has_scan", "has_member_day", "has_mall_buy", "has_visit"]].sum(axis=1)

    identity_nonempty = member["身份标签"].fillna("").astype(str).str.strip()
    identity_nonempty = identity_nonempty[identity_nonempty != ""]

    age_series = member["年龄"].dropna()
    age_bins = [16, 26, 31, 36, 41, 46, 51, 56, 71]
    age_labels = ["16-25", "26-30", "31-35", "36-40", "41-45", "46-50", "51-55", "56-70"]
    age_group = pd.cut(age_series, bins=age_bins, right=False, labels=age_labels)
    age_dist = (age_group.value_counts().reindex(age_labels, fill_value=0) / len(age_series) * 100).round(1)

    gender_labeled = member["性别"].dropna()
    male = int((gender_labeled == 1).sum())
    female = int((gender_labeled == 0).sum())
    gender_total = int(len(gender_labeled))

    controllable = store[(store["门店类型"].isin(["专卖店", "涂装服务中心"])) & (store["微信门店状态"] == "上线")].copy()
    controllable_codes = set(controllable["门店编号"].astype(str))
    code_to_type = {str(r["门店编号"]): r["门店类型"] for _, r in store.iterrows()}

    attached["挂靠门店代码"] = attached["挂靠门店代码"].astype(str)
    attached["门店类型"] = attached["挂靠门店代码"].map(code_to_type)
    attached_selected = attached[attached["挂靠门店代码"].isin(controllable_codes)].copy()
    attached_selected = attached_selected.sort_values("累计挂靠一物一码会员人数", ascending=False)

    jobs, areas = read_fixed_area_tables(path)
    job_rows = [row for row in jobs if row[0] != "合计"]
    pending_total = next((int(row[1]) for row in jobs if row[0] == "合计"), sum(int(row[1]) for row in job_rows))
    top_job_row = max(job_rows, key=lambda row: row[1]) if job_rows else ["", 0, 0, 0, 0, 0]

    consumers = member[member["has_store_buy"]].copy()
    type_summary = member.groupby("门店类型简").agg(
        total=("会员编码", "count"),
        buyers=("has_store_buy", "sum"),
        amount=("2025年门店消费金额（千元）", "sum"),
    )
    type_freq = consumers.groupby("门店类型简")["2025年门店消费天数"].mean()
    type_avg_amount = consumers.groupby("门店类型简")["2025年门店消费金额（千元）"].mean()

    by_level = consumers.groupby("等级名称").agg(
        buyers=("会员编码", "count"),
        freq=("2025年门店消费天数", "mean"),
        avg_amount=("2025年门店消费金额（千元）", "mean"),
        amount=("2025年门店消费金额（千元）", "sum"),
    )

    recent_month = consumers["2025最近一次消费月"].dropna().astype(int).astype(str).str[-2:]
    month_counts = recent_month.value_counts().reindex([f"{i:02d}" for i in range(1, 13)], fill_value=0)

    buyers = member[member["has_store_buy"]].copy()
    buyers["last_month"] = buyers["2025最近一次消费月"].fillna(0).astype(int)
    flow_loss_risk = (
        (buyers["last_month"] <= 202509)
        & (~buyers["has_scan"])
        & (~buyers["has_member_day"])
        & (~buyers["has_mall_buy"])
    )
    need_attention = buyers["last_month"].isin([202510, 202511]) & ((~buyers["has_member_day"]) | (~buyers["has_scan"]))

    svip = member[member["等级"] == 6].copy()
    svip_visited = int((svip["SVIP拜访情况（2为已拜访）"].fillna(0) == 2).sum())

    visit_store = member.groupby(["挂靠门店编码", "挂靠门店名称"]).agg(
        total=("会员编码", "count"),
        times=("2025年上门基检次数", "sum"),
        coupon_get=("基检券获得数量", "sum"),
        coupon_use=("基检券核销数量", "sum"),
        linked_amount=("基检券连带消费金额（元）", "sum"),
    ).reset_index()
    visit_store = visit_store[visit_store["times"] > 0].copy()
    recruit = store[["门店编号", "上门基检25年招募会员人数", "上门基检25年维护会员人数"]].copy()
    recruit["门店编号"] = recruit["门店编号"].astype(str)
    visit_store["挂靠门店编码"] = visit_store["挂靠门店编码"].astype(str)
    visit_store = visit_store.merge(recruit, left_on="挂靠门店编码", right_on="门店编号", how="left")
    visit_store["拓新人数"] = visit_store["上门基检25年招募会员人数"].fillna(0).astype(int)
    visit_store["维旧人数"] = visit_store["上门基检25年维护会员人数"].fillna(0).astype(int)
    visit_store = visit_store.sort_values(["times", "total"], ascending=[False, False])

    member_day_month = member_day_detail.iloc[1:, 3].astype(str).str.extract(r"年(\d+)月")[0].astype(float)
    member_day_month = member_day_month.value_counts().sort_index()

    member_day_type = member.groupby("门店类型控").agg(
        total=("会员编码", "count"),
        participants=("has_member_day", "sum"),
    )
    member_day_level = member.groupby("等级名称").agg(
        total=("会员编码", "count"),
        participants=("has_member_day", "sum"),
        car=("大奖中奖数量\n（汽车、电动车）", "sum"),
        item=("普通实物中奖数量\n（不含车&色卡&辅材）", "sum"),
        color=("色卡中奖数量", "sum"),
        material_prize=("辅材中奖数量", "sum"),
        material_use=("辅材核销数量", "sum"),
        coupon_prize=("券中奖数量", "sum"),
        coupon_use=("券核销数量", "sum"),
        linked_amount=("券连带消费金额（元）", "sum"),
    )

    segment_name = pd.Series("", index=member.index, dtype="object")
    high_active = (member["active_count"] >= 3) | ((member["等级"] == 6) & (member["has_store_buy"] | member["has_member_day"] | member["has_mall_buy"]))
    converting = (~member["has_store_buy"]) & (member["has_scan"] | member["has_member_day"] | member["has_visit"] | member["has_mall_buy"])
    consume_maintain = member["has_store_buy"] & (~high_active)
    silent = ~(member["has_store_buy"] | member["has_scan"] | member["has_member_day"] | member["has_mall_buy"] | member["has_visit"])

    for name, mask in [
        ("高活跃核心", high_active),
        ("待转化潜力", converting),
        ("消费待维系", consume_maintain),
        ("沉默会员", silent),
    ]:
        segment_name = segment_name.mask((segment_name == "") & mask, name)
    segment_name = segment_name.replace("", "其他")
    segment_counts = segment_name.value_counts()

    controllable["base_score"] = controllable["累计注册人数（含挂靠）"].rank(pct=True)
    controllable["low_conv_score"] = 1 - controllable["2025年消费率"].rank(pct=True)
    controllable["attach_score"] = controllable["一物一码来源挂靠人数"].rank(pct=True)
    controllable["mid_scan_score"] = controllable["其中：一物一码≥50包挂靠人数"].rank(pct=True)
    controllable["lottery_score"] = controllable["2025年参加会员日抽奖人数"].rank(pct=True)
    controllable["mall_gap_score"] = 1 - controllable["2025年会员商城有购买的人数"].rank(pct=True)
    controllable["opportunity_score"] = (
        0.25 * controllable["base_score"]
        + 0.20 * controllable["low_conv_score"]
        + 0.20 * controllable["attach_score"]
        + 0.15 * controllable["mid_scan_score"]
        + 0.10 * controllable["lottery_score"]
        + 0.10 * controllable["mall_gap_score"]
    ).round(4)
    controllable = controllable.sort_values("opportunity_score", ascending=False)

    def store_tags(row: pd.Series) -> list[str]:
        tags: list[str] = []
        if row["累计注册人数（含挂靠）"] >= 900:
            tags.append("高会员盘")
        if row["2025年消费率"] < 0.08:
            tags.append("低消费转化")
        if row["一物一码来源挂靠人数"] >= 400:
            tags.append("挂靠基础强")
        if row["2025年会员商城有购买的人数"] < 25:
            tags.append("线上渗透弱")
        if (row["上门基检25年招募会员人数"] + row["上门基检25年维护会员人数"]) > 0:
            tags.append("基检可运营")
        return tags

    consumption_store_rows = [
        {
            "store_type": key,
            "total": int(type_summary.loc[key, "total"]),
            "buyers": int(type_summary.loc[key, "buyers"]),
            "buyer_pct": round(type_summary.loc[key, "buyers"] / type_summary.loc[key, "total"] * 100, 1)
            if type_summary.loc[key, "total"]
            else 0,
            "avg_freq": round(float(type_freq.get(key, 0)), 1),
            "avg_amount_k": round(float(type_avg_amount.get(key, 0)), 1),
            "amount_k": round(float(type_summary.loc[key, "amount"]), 1),
        }
        for key in ["专卖店", "涂装服务中心", "其他店型"]
    ]
    consumption_store_rows.append(
        {
            "store_type": "合计",
            "total": int(len(member)),
            "buyers": int(member["has_store_buy"].sum()),
            "buyer_pct": round(member["has_store_buy"].sum() / len(member) * 100, 1) if len(member) else 0,
            "avg_freq": round(float(consumers["2025年门店消费天数"].mean()), 1) if len(consumers) else 0,
            "avg_amount_k": round(float(consumers["2025年门店消费金额（千元）"].mean()), 1) if len(consumers) else 0,
            "amount_k": round(float(consumers["2025年门店消费金额（千元）"].sum()), 1),
        }
    )

    consumption_level_rows = [
        {
            "level": key,
            "buyers": int(by_level.loc[key, "buyers"]),
            "avg_freq": round(float(by_level.loc[key, "freq"]), 1),
            "avg_amount_k": round(float(by_level.loc[key, "avg_amount"]), 1),
            "amount_k": round(float(by_level.loc[key, "amount"]), 1),
            "buyer_pct": round(by_level.loc[key, "buyers"] / len(consumers) * 100, 1) if len(consumers) else 0,
            "amount_pct": round(by_level.loc[key, "amount"] / consumers["2025年门店消费金额（千元）"].sum() * 100, 1)
            if consumers["2025年门店消费金额（千元）"].sum()
            else 0,
        }
        for key in LEVEL_ORDER
    ]
    consumption_level_rows.append(
        {
            "level": "合计",
            "buyers": int(len(consumers)),
            "avg_freq": round(float(consumers["2025年门店消费天数"].mean()), 1) if len(consumers) else 0,
            "avg_amount_k": round(float(consumers["2025年门店消费金额（千元）"].mean()), 1) if len(consumers) else 0,
            "amount_k": round(float(consumers["2025年门店消费金额（千元）"].sum()), 1),
            "buyer_pct": 100.0 if len(consumers) else 0,
            "amount_pct": 100.0 if len(consumers) else 0,
        }
    )

    visit_rows = [
        {
            "store_code": row["挂靠门店编码"],
            "store_name": row["挂靠门店名称"],
            "covered_members": int(row["total"]),
            "times": int(row["times"]),
            "maintain": int(row["维旧人数"]),
            "recruit": int(row["拓新人数"]),
            "coupon_get": int(row["coupon_get"]),
            "coupon_use": int(row["coupon_use"]),
            "coupon_pct": round(row["coupon_use"] / row["coupon_get"] * 100, 1) if row["coupon_get"] else 0,
            "linked_amount": int(row["linked_amount"]),
        }
        for _, row in visit_store.iterrows()
    ]
    visit_rows.append(
        {
            "store_code": "合计",
            "store_name": "合计",
            "covered_members": int(visit_store["total"].sum()),
            "times": int(visit_store["times"].sum()),
            "maintain": int(visit_store["维旧人数"].sum()),
            "recruit": int(visit_store["拓新人数"].sum()),
            "coupon_get": int(visit_store["coupon_get"].sum()),
            "coupon_use": int(visit_store["coupon_use"].sum()),
            "coupon_pct": round(visit_store["coupon_use"].sum() / visit_store["coupon_get"].sum() * 100, 1)
            if visit_store["coupon_get"].sum()
            else 0,
            "linked_amount": int(visit_store["linked_amount"].sum()),
        }
    )

    member_day_store_rows = [
        {
            "store_type": key,
            "participants": int(member_day_type.loc[key, "participants"]),
            "pct": round(member_day_type.loc[key, "participants"] / member_day_type.loc[key, "total"] * 100, 1)
            if member_day_type.loc[key, "total"]
            else 0,
        }
        for key in ["专卖店", "涂装服务中心", "非可控店"]
    ]
    member_day_store_rows.append(
        {
            "store_type": "合计",
            "participants": int(member["has_member_day"].sum()),
            "pct": round(member["has_member_day"].sum() / len(member) * 100, 1) if len(member) else 0,
        }
    )

    member_day_level_rows = [
        {
            "level": key,
            "total": int(member_day_level.loc[key, "total"]),
            "participants": int(member_day_level.loc[key, "participants"]),
            "participant_pct": round(member_day_level.loc[key, "participants"] / member_day_level.loc[key, "total"] * 100, 1)
            if member_day_level.loc[key, "total"]
            else 0,
            "car": int(member_day_level.loc[key, "car"]),
            "item": int(member_day_level.loc[key, "item"]),
            "color": int(member_day_level.loc[key, "color"]),
            "material_prize": int(member_day_level.loc[key, "material_prize"]),
            "material_use": int(member_day_level.loc[key, "material_use"]),
            "coupon_prize": int(member_day_level.loc[key, "coupon_prize"]),
            "coupon_use": int(member_day_level.loc[key, "coupon_use"]),
            "linked_amount": int(member_day_level.loc[key, "linked_amount"]),
        }
        for key in LEVEL_ORDER
    ]
    member_day_level_rows.append(
        {
            "level": "合计",
            "total": int(member_day_level["total"].sum()),
            "participants": int(member_day_level["participants"].sum()),
            "participant_pct": round(member_day_level["participants"].sum() / member_day_level["total"].sum() * 100, 1)
            if member_day_level["total"].sum()
            else 0,
            "car": int(member_day_level["car"].sum()),
            "item": int(member_day_level["item"].sum()),
            "color": int(member_day_level["color"].sum()),
            "material_prize": int(member_day_level["material_prize"].sum()),
            "material_use": int(member_day_level["material_use"].sum()),
            "coupon_prize": int(member_day_level["coupon_prize"].sum()),
            "coupon_use": int(member_day_level["coupon_use"].sum()),
            "linked_amount": int(member_day_level["linked_amount"].sum()),
        }
    )

    analysis = {
        "meta": {
            "input": str(path),
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "template_mode": "fixed_member_ops_excel",
        },
        "workbook_summary": {
            "sheet_names": ["全部会员", "门店别", "一物一码来源已挂靠会员分析", "扫码>=50游离会员", "每月会员日参与名单"],
            "member_rows": int(len(member)),
            "store_rows": int(len(store)),
            "controllable_online_stores": int(len(controllable)),
        },
        "pages": {
            "member_overview": {
                "total_members": int(len(member)),
                "identity_labeled": int(len(identity_nonempty)),
                "gender_labeled": gender_total,
                "age_labeled": int(len(age_series)),
                "avg_age": round(float(age_series.mean()), 1),
                "male": male,
                "female": female,
                "male_pct": round(male / gender_total * 100, 1) if gender_total else 0,
                "female_pct": round(female / gender_total * 100, 1) if gender_total else 0,
                "identity_top": [
                    {
                        "name": name,
                        "value": int(value),
                        "pct": round(value / len(identity_nonempty) * 100, 1),
                    }
                    for name, value in identity_nonempty.value_counts().head(10).items()
                ],
                "age_distribution": [{"label": label, "pct": float(age_dist[label])} for label in age_labels],
                "level_distribution": [
                    {
                        "name": LEVEL_MAP[level],
                        "value": int((member["等级"] == level).sum()),
                    }
                    for level in sorted(LEVEL_MAP)
                ],
                "qr_attached_members": int(store["一物一码来源挂靠人数"].fillna(0).sum()),
            },
            "qr_attached": {
                "summary": {
                    "store_count": int(len(attached_selected)),
                    "attached_members": int(attached_selected["累计挂靠会员人数"].sum()),
                    "qr_attached_members": int(attached_selected["累计挂靠一物一码会员人数"].sum()),
                    "qr_attached_pct": round(
                        float(attached_selected["累计挂靠一物一码会员人数"].sum() / attached_selected["累计挂靠会员人数"].sum() * 100), 1
                    )
                    if attached_selected["累计挂靠会员人数"].sum()
                    else 0,
                    "lottery_members": int(attached_selected["累计挂靠一物一码会员2025年有参加会员日抽奖人数"].sum()),
                    "redeem_members": int(attached_selected["累计挂靠一物一码会员2025年有核销产品礼包/券人数"].sum()),
                },
                "rows": [
                    {
                        "store_code": row["挂靠门店代码"],
                        "store_name": row["挂靠门店名称"],
                        "store_type": row["门店类型"],
                        "attached_members": int(row["累计挂靠会员人数"]),
                        "qr_attached_members": int(row["累计挂靠一物一码会员人数"]),
                        "qr_attached_pct": round(float(row["累计挂靠一物一码会员人数"] / row["累计挂靠会员人数"] * 100), 1)
                        if row["累计挂靠会员人数"]
                        else 0,
                        "buyers": int(row["累计挂靠会员2025年有消费人数"]),
                        "buyers_amount_k": round(float(row["累计挂靠会员2025年消费A价（未税 千元）"]), 0),
                        "qr_buyers": int(row["累计挂靠一物一码会员2025年有消费人数"]),
                        "qr_buyers_amount_k": round(float(row["累计挂靠一物一码会员2025年消费A价（未税 千元）"]), 0),
                        "lottery_members": int(row["累计挂靠一物一码会员2025年有参加会员日抽奖人数"]),
                        "redeem_members": int(row["累计挂靠一物一码会员2025年有核销产品礼包/券人数"]),
                    }
                    for _, row in attached_selected.iterrows()
                ],
            },
            "qr_pending": {
                "summary": {
                    "pending_total": int(pending_total),
                    "top_job": top_job_row[0],
                    "top_job_pct": round(top_job_row[1] / pending_total * 100, 1) if pending_total else 0,
                },
                "job_table": jobs,
                "area_table": areas,
            },
            "consumption": {
                "overview": {
                    "total_members": int(len(member)),
                    "store_buyers": int(member["has_store_buy"].sum()),
                    "scan_buyers": int(member["has_scan"].sum()),
                    "both_buyers": int((member["has_store_buy"] & member["has_scan"]).sum()),
                    "union_members": int(member["has_union"].sum()),
                },
                "by_store_type": consumption_store_rows,
                "by_level": consumption_level_rows,
                "recent_consume_months": [
                    {"month": f"{i}月", "value": int(month_counts[f"{i:02d}"])}
                    for i in range(1, 13)
                ],
                "risk_pool": {
                    "flow_loss_risk": int(flow_loss_risk.sum()),
                    "need_attention": int(need_attention.sum()),
                },
                "svip": {
                    "total": int(len(svip)),
                    "visited": svip_visited,
                    "visit_pct": round(svip_visited / len(svip) * 100, 1) if len(svip) else 0,
                },
            },
            "home_visit": {
                "summary": {
                    "covered_members": int(visit_store["total"].sum()),
                    "times": int(visit_store["times"].sum()),
                    "maintain": int(visit_store["维旧人数"].sum()),
                    "recruit": int(visit_store["拓新人数"].sum()),
                    "coupon_get": int(visit_store["coupon_get"].sum()),
                    "coupon_use": int(visit_store["coupon_use"].sum()),
                    "coupon_pct": round(visit_store["coupon_use"].sum() / visit_store["coupon_get"].sum() * 100, 1)
                    if visit_store["coupon_get"].sum()
                    else 0,
                    "linked_amount": int(visit_store["linked_amount"].sum()),
                },
                "rows": visit_rows,
            },
            "member_day": {
                "summary": {
                    "participants": int(member["has_member_day"].sum()),
                    "participant_pct": round(member["has_member_day"].sum() / len(member) * 100, 1) if len(member) else 0,
                    "avg_times": round(float(member.loc[member["has_member_day"], "2025年参加会员日次数"].mean()), 1)
                    if member["has_member_day"].sum()
                    else 0,
                    "svip_pct": round(member_day_level.loc["SVIP", "participants"] / member_day_level.loc["SVIP", "total"] * 100, 1)
                    if "SVIP" in member_day_level.index
                    else 0,
                },
                "by_store_type": member_day_store_rows,
                "month_trend": [{"month": f"{int(key)}月", "value": int(value)} for key, value in member_day_month.items()],
                "times_distribution": [
                    {
                        "times": int(times),
                        "members": int(count),
                        "pct": round(count / member["has_member_day"].sum() * 100, 1) if member["has_member_day"].sum() else 0,
                    }
                    for times, count in member.loc[member["has_member_day"], "2025年参加会员日次数"].astype(int).value_counts().sort_index(ascending=False).items()
                ],
                "by_level": member_day_level_rows,
            },
            "segment_risk": {
                "segments": [
                    {
                        "name": name,
                        "members": int(segment_counts.get(name, 0)),
                        "pct": round(segment_counts.get(name, 0) / len(member) * 100, 1),
                    }
                    for name in ["高活跃核心", "消费待维系", "待转化潜力", "沉默会员"]
                ],
                "risk_rules": {
                    "flow_loss_risk": "历史有消费，最近消费月份较早，且近期无扫码、无会员日、无商城行为",
                    "need_attention": "最近消费集中在10-11月，但互动行为不足",
                },
            },
            "store_priority": {
                "rows": [
                    {
                        "store_code": str(row["门店编号"]),
                        "store_name": row["门店名称"],
                        "store_type": row["门店类型"],
                        "member_base": int(row["累计注册人数（含挂靠）"]),
                        "consume_pct": round(float(row["2025年消费率"]) * 100, 1),
                        "qr_attach": int(row["一物一码来源挂靠人数"]),
                        "mall_buyers": int(row["2025年会员商城有购买的人数"]),
                        "opportunity_score": round(float(row["opportunity_score"]) * 100, 1),
                        "tags": store_tags(row),
                    }
                    for _, row in controllable.iterrows()
                ]
            },
            "conversion_path": {
                "touchpoints": [
                    {"name": "会员总盘", "value": int(len(member))},
                    {"name": "身份已标注", "value": int(len(identity_nonempty))},
                    {"name": "一物一码来源", "value": int((member["是否一物一码来源"].fillna("").astype(str).str.strip() != "").sum())},
                    {"name": "有扫码行为", "value": int(member["has_scan"].sum())},
                    {"name": "会员日参与", "value": int(member["has_member_day"].sum())},
                    {"name": "门店消费", "value": int(member["has_store_buy"].sum())},
                    {"name": "商城购买", "value": int(member["has_mall_buy"].sum())},
                    {"name": "SVIP", "value": int((member["等级"] == 6).sum())},
                ]
            },
        },
    }

    return analysis


def main() -> None:
    args = parse_args()
    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    result = build_analysis(input_path)
    output_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(output_path)


if __name__ == "__main__":
    main()
