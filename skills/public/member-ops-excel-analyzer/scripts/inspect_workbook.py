#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


REQUIRED_SHEETS = [
    "全部会员",
    "门店别",
    "一物一码来源已挂靠会员分析",
    "扫码>=50游离会员",
    "每月会员日参与名单",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Inspect fixed member-ops workbook template.")
    parser.add_argument("--input", required=True, help="Path to the Excel workbook.")
    parser.add_argument("--output", help="Optional path to save inspection JSON.")
    return parser.parse_args()


def workbook_summary(path: Path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_summaries = []
    for ws in wb.worksheets:
      sheet_summaries.append(
          {
              "title": ws.title,
              "rows": ws.max_row,
              "cols": ws.max_column,
          }
      )

    missing = [name for name in REQUIRED_SHEETS if name not in wb.sheetnames]
    return {
        "input": str(path),
        "sheet_names": wb.sheetnames,
        "required_sheets_present": len(missing) == 0,
        "missing_required_sheets": missing,
        "sheet_summaries": sheet_summaries,
    }


def main() -> None:
    args = parse_args()
    path = Path(args.input).expanduser().resolve()
    summary = workbook_summary(path)
    payload = json.dumps(summary, ensure_ascii=False, indent=2)
    print(payload)
    if args.output:
        out = Path(args.output).expanduser().resolve()
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(payload, encoding="utf-8")


if __name__ == "__main__":
    main()
